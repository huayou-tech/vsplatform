"""
================================================================
微信公众号 + CSDN + 知乎 多平台发布计划生成器
功能：
1. 生成 Excel 格式的 7 天发布计划
2. 为每天的内容自动生成三个平台版本
3. 支持数据追踪和统计

作者：华又科技团队
版本：V1.0
创建时间：2026-03-11
================================================================
"""

import os
import json
import pandas as pd
from datetime import datetime, timedelta
import re


class MultiPlatformPublisher:
    """多平台发布管理器"""
    
    def __init__(self, workspace_path):
        self.workspace = workspace_path
        self.articles_dir = os.path.join(workspace_path, "CSDN_ARTICLES")
        self.output_dir = os.path.join(workspace_path, "multi_platform_output")
        self.publish_log_path = os.path.join(workspace_path, "publish_log.json")
        
        # 创建输出目录
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        
        # 加载已发布记录
        self.published_articles = self._load_publish_log()
        
        # 待发布的文章列表
        self.pending_articles = self._scan_pending_articles()
    
    def _load_publish_log(self):
        """加载发布日志"""
        if os.path.exists(self.publish_log_path):
            with open(self.publish_log_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('published_articles', [])
        return []
    
    def _scan_pending_articles(self):
        """扫描待发布的文章"""
        pending = []
        published_files = [article['file_path'] for article in self.published_articles]
        
        # 支持的扩展名
        extensions = ['.md', '.txt']
        
        for filename in os.listdir(self.articles_dir):
            filepath = os.path.join(self.articles_dir, filename)
            
            # 只处理文件，跳过目录
            if not os.path.isfile(filepath):
                continue
            
            # 检查扩展名
            if not any(filename.endswith(ext) for ext in extensions):
                continue
            
            # 跳过 README 等说明文件
            if filename.startswith('README'):
                continue
            
            # 跳过已发布的
            relative_path = os.path.join("CSDN_ARTICLES", filename)
            if relative_path in published_files:
                continue
            
            # 提取文章信息
            article_info = {
                'filename': filename,
                'filepath': filepath,
                'relative_path': relative_path,
                'title': self._extract_title(filepath),
                'size': os.path.getsize(filepath)
            }
            pending.append(article_info)
        
        # 按文件名排序
        pending.sort(key=lambda x: x['filename'])
        
        return pending
    
    def _extract_title(self, filepath):
        """从文件中提取标题"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read(500)  # 只读前 500 字符
                
                # 尝试提取 Markdown 标题
                lines = content.split('\n')
                for line in lines:
                    if line.startswith('# '):
                        return line[2:].strip()
                
                # 如果没有 Markdown 标题，用文件名
                filename = os.path.basename(filepath)
                title = os.path.splitext(filename)[0]
                return title
        except Exception as e:
            print(f"读取文件失败：{e}")
            return "未知标题"
    
    def generate_excel_plan(self, days=7, start_date=None):
        """生成 Excel 发布计划表"""
        if start_date is None:
            start_date = datetime.now()
        
        # 如果待发布文章不够，循环使用
        articles = self.pending_articles.copy()
        if len(articles) == 0:
            print("警告：没有待发布的文章！")
            return None
        
        # 扩展文章列表（如果不够 7 天）
        while len(articles) < days:
            articles.extend(articles)
        articles = articles[:days]
        
        # 创建计划数据
        plan_data = []
        for i in range(days):
            date = start_date + timedelta(days=i)
            article = articles[i]
            
            day_plan = {
                '日期': date.strftime('%Y-%m-%d'),
                '星期': ['周一', '周二', '周三', '周四', '周五', '周六', '周日'][date.weekday()],
                '第几天': f'Day {i+1}',
                '文章标题': article['title'],
                '源文件': article['filename'],
                
                # CSDN 平台
                'CSDN_发布时间': '08:30',
                'CSDN_标题风格': '技术型',
                'CSDN_分类': 'C++',
                'CSDN_标签': 'C++,视频监控，网络编程',
                'CSDN_状态': '待发布',
                
                # 知乎平台
                '知乎_发布时间': '12:00',
                '知乎_回答话题': 'SIP协议，GB/T 28181,视频监控',
                '知乎_状态': '待发布',
                
                # 公众号平台
                '公众号_发布时间': '20:00',
                '公众号_标题风格': '吸引眼球型',
                '公众号_封面图': '待准备',
                '公众号_状态': '待发布',
                
                # 社群运营
                '社群分享时间': '21:00',
                '分享群数': 3,
                '预计阅读量': 100 * (i + 1),
                
                # 数据追踪
                '实际阅读量': '',
                '评论数': '',
                '新增粉丝': '',
                '咨询数': '',
                '收入': '',
                '备注': ''
            }
            plan_data.append(day_plan)
        
        # 创建 DataFrame
        df = pd.DataFrame(plan_data)
        
        # 保存为 Excel
        excel_path = os.path.join(self.output_dir, f'发布计划_{start_date.strftime("%Y%m%d")}.xlsx')
        
        # 使用 ExcelWriter 设置格式
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='7 天发布计划')
            
            # 获取 workbook 和 worksheet
            workbook = writer.book
            worksheet = writer.sheets['7 天发布计划']
            
            # 设置列宽
            column_widths = {
                'A': 12,  # 日期
                'B': 8,   # 星期
                'C': 10,  # 第几天
                'D': 40,  # 文章标题
                'E': 25,  # 源文件
                'F': 15,  # CSDN_发布时间
                'G': 12,  # CSDN_标题风格
                'H': 15,  # CSDN_分类
                'I': 30,  # CSDN_标签
                'J': 12,  # CSDN_状态
                'K': 15,  # 知乎_发布时间
                'L': 35,  # 知乎_回答话题
                'M': 12,  # 知乎_状态
                'N': 15,  # 公众号_发布时间
                'O': 15,  # 公众号_标题风格
                'P': 15,  # 公众号_封面图
                'Q': 12,  # 公众号_状态
                'R': 15,  # 社群分享时间
                'S': 12,  # 分享群数
                'T': 15,  # 预计阅读量
                'U': 12,  # 实际阅读量
                'V': 12,  # 评论数
                'W': 12,  # 新增粉丝
                'X': 12,  # 咨询数
                'Y': 12,  # 收入
                'Z': 30,  # 备注
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置行高
            for row in worksheet.iter_rows(min_row=1, max_row=len(plan_data) + 1):
                worksheet.row_dimensions[row[0].row].height = 25
            
            # 设置标题行样式（加粗、背景色）
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 设置数据区域样式
            data_alignment = Alignment(horizontal="left", vertical="center")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            for row in worksheet.iter_rows(min_row=2, max_row=len(plan_data) + 1):
                for cell in row:
                    cell.alignment = data_alignment
                    # 状态列居中
                    if cell.column_letter in ['J', 'M', 'Q']:
                        cell.alignment = center_alignment
            
            # 添加条件格式（状态列）
            from openpyxl.formatting.rule import FormulaRule
            from openpyxl.styles import Font
            
            # 待发布 - 橙色
            for col in ['J', 'M', 'Q']:
                worksheet.conditional_formatting.add(
                    f'{col}2:{col}{len(plan_data) + 1}',
                    FormulaRule(
                        formula=[f'${col}2="待发布"'],
                        font=Font(color="FF6600", bold=True)
                    )
                )
            
            # 已发布 - 绿色
            for col in ['J', 'M', 'Q']:
                worksheet.conditional_formatting.add(
                    f'{col}2:{col}{len(plan_data) + 1}',
                    FormulaRule(
                        formula=[f'${col}2="已发布"'],
                        font=Font(color="00AA00", bold=True)
                    )
                )
        
        print(f"✅ Excel 发布计划已生成：{excel_path}")
        return excel_path
    
    def convert_to_csdn_format(self, source_content, title):
        """转换为 CSDN 格式（TXT）"""
        # 清理 Markdown 格式
        content = source_content
        
        # 移除 Markdown 标题标记
        content = re.sub(r'^#+\s*', '', content, flags=re.MULTILINE)
        
        # 移除代码块标记但保留内容
        content = re.sub(r'```\w*\n', '\n', content)
        content = re.sub(r'\n```', '\n', content)
        
        # 移除列表标记
        content = re.sub(r'^[\-\*]\s*', '', content, flags=re.MULTILINE)
        content = re.sub(r'^\d+\.\s*', '', content, flags=re.MULTILINE)
        
        # 移除引用标记
        content = re.sub(r'^>\s*', '', content, flags=re.MULTILINE)
        
        # 移除粗体/斜体标记
        content = re.sub(r'\*\*(.+?)\*\*', r'\1', content)
        content = re.sub(r'\*(.+?)\*', r'\1', content)
        
        # 删除连续多余空行（只保留单个空行）
        content = re.sub(r'\n{3,}', '\n\n', content)
        
        # 移除行首尾空白
        lines = [line.strip() for line in content.split('\n')]
        content = '\n'.join(lines)
        
        # 添加导读信息
        output = f"""================================================================
导读：本文是华又科技视频监控平台系列技术文章之一。我们开源了完整的视频监控系统源码，包含 23 个核心工程项目。
GitHub 地址：https://github.com/huayou-tech/videomonitor-platform
技术栈：C++17 | Windows IOCP | Live555 | SIP/GB28181 | RTSP/ONVIF
================================================================

{title}

{content}

================================================================
关于项目
================================================================
本项目是一个完整的商业级视频监控平台，经过十年验证，稳定可靠。

资源链接：
- GitHub 源码：https://github.com/huayou-tech/videomonitor-platform
- 技术博客：关注作者获取更多系列文章

支持作者：
- Star 我们的 GitHub 项目
- 点赞 + 收藏 + 评论
- 分享给更多需要的人
"""
        
        return output
    
    def convert_to_zhihu_format(self, source_content, title):
        """转换为知乎格式（Markdown，增强可读性）"""
        content = source_content
        
        # 知乎喜欢更结构化的内容
        # 在开头添加导语
        intro = f"""**导语：** 这是我在视频监控行业深耕 10 年的经验总结。今天把这份实战经验分享给大家，希望能帮你少走弯路。

---

"""
        
        # 优化标题层级
        content = re.sub(r'^####\s*', '### ', content, flags=re.MULTILINE)
        content = re.sub(r'^#####\s*', '#### ', content, flags=re.MULTILINE)
        
        # 添加更多分段符
        content = content.replace('\n\n', '\n\n---\n\n')
        
        # 在代码块前添加说明
        content = re.sub(r'(```\w+)', r'\n【代码示例】\n\1', content)
        
        output = intro + content
        
        # 添加结尾互动
        ending = f"""
---

**写在最后：**

如果你在开发过程中遇到类似问题，欢迎在评论区留言讨论！

👍 **觉得有帮助？** 请点个赞，让更多人看到！

📚 **延伸阅读：** 我的其他技术文章：
- Windows IOCP异步模型深度解析
- 通用函数库 CommonFunc 封装
- 设备状态监控服务实现

💼 **技术服务：**
- 技术咨询：39.9 元/次
- 远程指导：99 元/小时
- 定制开发：根据需求报价

联系方式见个人主页！
"""
        
        output += ending
        return output
    
    def convert_to_wechat_format(self, source_content, title):
        """转换为公众号格式（HTML 片段，可用于秀米）"""
        content = source_content
        
        # 公众号文章结构
        html_template = """<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
</head>
<body style="margin: 0; padding: 20px; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;">
    
    <!-- 导语 -->
    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 10px; margin-bottom: 30px; text-align: center;">
        <p style="font-size: 16px; margin: 0; line-height: 1.6;">
            💡 <strong>做视频监控 10 年，踩过最多的坑就是{keywords}。今天把这份血泪总结分享给大家，希望能帮你少走弯路。</strong>
        </p>
    </div>
    
    <!-- 正文内容 -->
    <div style="font-size: 16px; line-height: 1.75; color: #333;">
        {content}
    </div>
    
    <!-- 福利区 -->
    <div style="background: #f0f9ff; border-left: 4px solid #0066CC; padding: 20px; margin-top: 30px; border-radius: 5px;">
        <p style="font-size: 16px; color: #0066CC; margin: 0 0 10px 0;">
            🎁 <strong>限时福利</strong>
        </p>
        <ul style="color: #333; padding-left: 20px;">
            <li>回复"sip"获取完整源码</li>
            <li>前 10 名咨询送 30 分钟免费指导</li>
            <li>加入技术交流群</li>
        </ul>
    </div>
    
    <!-- 联系方式 -->
    <div style="text-align: center; margin-top: 30px; padding: 20px; background: #f5f5f5; border-radius: 10px;">
        <p style="font-size: 14px; color: #666; margin: 5px 0;">
            <strong>华又科技 - 专注视频监控技术开发 10 年</strong>
        </p>
        <p style="font-size: 14px; color: #999; margin: 5px 0;">
            提供：技术咨询 | 定制开发 | 远程指导
        </p>
        <p style="font-size: 14px; color: #0066CC; margin: 10px 0;">
            公众号：华又科技
        </p>
    </div>
    
</body>
</html>
"""
        
        # 提取关键词（从标题）
        keywords = "SIP协议" if "SIP" in title else "视频监控"
        
        # 转换内容为 HTML 段落
        paragraphs = content.split('\n\n')
        html_content = ""
        
        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
            
            # 如果是代码块
            if para.startswith('```'):
                code_lines = para.split('\n')[1:-1]  # 移除 ``` 标记
                code_html = '<pre style="background: #f5f5f5; padding: 15px; border-radius: 5px; overflow-x: auto; border-left: 4px solid #0066CC;">'
                for line in code_lines:
                    code_html += f'<code style="font-family: Consolas, monospace; font-size: 14px;">{line}</code><br>'
                code_html += '</pre>'
                html_content += code_html + '\n\n'
            
            # 如果是标题（以#开头）
            elif para.startswith('#'):
                title_text = re.sub(r'^#+\s*', '', para)
                html_content += f'<h2 style="color: #0066CC; border-bottom: 2px solid #0066CC; padding-bottom: 10px; margin-top: 30px;">{title_text}</h2>\n\n'
            
            # 如果是列表
            elif para.startswith('- ') or para.startswith('* '):
                items = [re.sub(r'^[-*]\s*', '', line.strip()) for line in para.split('\n') if line.strip().startswith('- ') or line.strip().startswith('* ')]
                html_content += '<ul style="background: #f9f9f9; padding: 15px 35px; border-radius: 5px;">\n'
                for item in items:
                    html_content += f'<li style="margin: 8px 0;">{item}</li>\n'
                html_content += '</ul>\n\n'
            
            # 普通段落
            else:
                # 处理强调内容
                para = re.sub(r'\*\*(.+?)\*\*', r'<strong style="color: #0066CC;">\1</strong>', para)
                para = re.sub(r'`(.+?)`', r'<code style="background: #f0f0f0; padding: 2px 6px; border-radius: 3px; font-family: Consolas, monospace;">\1</code>', para)
                
                html_content += f'<p style="margin: 15px 0; text-align: justify;">{para}</p>\n\n'
        
        # 填充模板
        final_html = html_template.format(
            title=title,
            keywords=keywords,
            content=html_content
        )
        
        return final_html
    
    def generate_daily_content(self, article_info, day_num):
        """为一篇文章生成三个平台版本"""
        # 读取源文件
        with open(article_info['filepath'], 'r', encoding='utf-8') as f:
            source_content = f.read()
        
        title = article_info['title']
        
        # 生成三个平台的版本
        csdn_content = self.convert_to_csdn_format(source_content, title)
        zhihu_content = self.convert_to_zhihu_format(source_content, title)
        wechat_content = self.convert_to_wechat_format(source_content, title)
        
        # 保存文件
        output_date = (datetime.now() + timedelta(days=day_num-1)).strftime('%Y%m%d')
        
        # CSDN 版本
        csdn_filename = f"{output_date}_Day{day_num}_CSDN.txt"
        csdn_path = os.path.join(self.output_dir, csdn_filename)
        with open(csdn_path, 'w', encoding='utf-8') as f:
            f.write(csdn_content)
        
        # 知乎版本
        zhihu_filename = f"{output_date}_Day{day_num}_Zhihu.md"
        zhihu_path = os.path.join(self.output_dir, zhihu_filename)
        with open(zhihu_path, 'w', encoding='utf-8') as f:
            f.write(zhihu_content)
        
        # 公众号版本
        wechat_filename = f"{output_date}_Day{day_num}_WeChat.html"
        wechat_path = os.path.join(self.output_dir, wechat_filename)
        with open(wechat_path, 'w', encoding='utf-8') as f:
            f.write(wechat_content)
        
        print(f"✅ Day {day_num} 内容已生成:")
        print(f"   - CSDN: {csdn_filename}")
        print(f"   - 知乎：{zhihu_filename}")
        print(f"   - 公众号：{wechat_filename}")
        
        return {
            'csdn': csdn_path,
            'zhihu': zhihu_path,
            'wechat': wechat_path
        }
    
    def run_full_generation(self, days=7):
        """执行完整的生成流程"""
        print("=" * 60)
        print("多平台发布内容生成器")
        print("=" * 60)
        print(f"工作目录：{self.workspace}")
        print(f"待发布文章：{len(self.pending_articles)}篇")
        print(f"计划生成：{days}天的内容")
        print("=" * 60)
        
        # 步骤 1: 生成 Excel 计划
        print("\n【步骤 1】生成 Excel 发布计划...")
        excel_path = self.generate_excel_plan(days=days)
        
        # 步骤 2: 为每天生成三平台内容
        print(f"\n【步骤 2】生成{days}天的多平台内容...")
        generated_files = []
        
        for i in range(days):
            article_idx = i % len(self.pending_articles)
            article = self.pending_articles[article_idx]
            
            print(f"\n>>> Day {i+1}: 《{article['title']}》")
            files = self.generate_daily_content(article, i+1)
            generated_files.append(files)
        
        # 步骤 3: 生成汇总报告
        print(f"\n【步骤 3】生成汇总报告...")
        summary_path = self._generate_summary_report(generated_files)
        
        print("\n" + "=" * 60)
        print("✅ 全部完成！")
        print("=" * 60)
        print(f"📊 Excel 计划表：{excel_path}")
        print(f"📝 内容文件：{self.output_dir}")
        print(f"📋 汇总报告：{summary_path}")
        print("=" * 60)
        
        return {
            'excel': excel_path,
            'files': generated_files,
            'summary': summary_path
        }
    
    def _generate_summary_report(self, generated_files):
        """生成汇总报告"""
        report = """================================================================
多平台内容生成汇总报告
================================================================

生成时间：{time}
总计：{total_days}天
输出目录：{output_dir}

================================================================
每日内容清单
================================================================

""".format(
            time=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            total_days=len(generated_files),
            output_dir=self.output_dir
        )
        
        for i, files in enumerate(generated_files, 1):
            report += f"""Day {i}:
  📄 CSDN 版本：{os.path.basename(files['csdn'])}
     - 发布时间：08:30
     - 格式：TXT（便于复制到 CSDN 编辑器）
  
  📝 知乎版本：{os.path.basename(files['zhihu'])}
     - 发布时间：12:00
     - 格式：Markdown（适配知乎排版）
  
  💬 公众号版本：{os.path.basename(files['wechat'])}
     - 发布时间：20:00
     - 格式：HTML（可导入秀米编辑器）

"""
        
        report += """================================================================
使用说明
================================================================

1. CSDN 发布：
   - 打开对应的 TXT 文件
   - 复制全部内容
   - 粘贴到 CSDN 编辑器
   - 设置分类和标签后发布

2. 知乎发布：
   - 打开对应的 MD 文件
   - 复制内容
   - 在知乎写回答/文章
   - 粘贴内容（支持 Markdown）

3. 公众号发布：
   - 打开对应的 HTML 文件
   - 用浏览器打开查看效果
   - 复制内容到秀米编辑器
   - 调整细节后同步到公众号

================================================================
注意事项
================================================================

1. 发布时间建议严格遵守（培养用户习惯）
2. 三个平台的内容要同一天发布（最大化曝光）
3. 记得及时回复评论和私信
4. 每天睡前检查数据并记录到 Excel

================================================================
祝发布顺利！
================================================================
"""
        
        # 保存报告
        report_path = os.path.join(self.output_dir, f'生成汇总报告_{datetime.now().strftime("%Y%m%d")}.txt')
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report)
        
        return report_path


# ============================================================================
# 主函数
# ============================================================================

if __name__ == '__main__':
    # 工作空间路径
    workspace = r'e:\华又科技\VSS\VSPlatform'
    
    # 创建发布管理器
    publisher = MultiPlatformPublisher(workspace)
    
    # 执行完整生成流程（生成 7 天内容）
    result = publisher.run_full_generation(days=7)
    
    print("\n\n下一步操作：")
    print("1. 打开 Excel 表格查看发布计划")
    print(f"   文件位置：{result['excel']}")
    print("\n2. 查看生成的文章内容")
    print(f"   文件夹：{result['files'][0]['csdn'].rsplit(os.sep, 1)[0]}")
    print("\n3. 按照计划每天发布内容")
    print("   记得回复评论和统计数据哦！")
